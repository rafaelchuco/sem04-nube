from __future__ import annotations

import atexit
import os
import re
from http import HTTPStatus
from io import BytesIO

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from onpe_service import (
    CaptchaRequiredError,
    OnpeLookupError,
    OnpePlaywrightClient,
)

app = Flask(__name__)


def _env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "si", "on"}


# Reutiliza la misma instancia de navegador para evitar el costo de abrir Chromium por request.
client = OnpePlaywrightClient(headless=_env_bool("ONPE_HEADLESS", default=False))


def _validar_dni(dni: str) -> bool:
    return bool(re.fullmatch(r"\d{8}", dni))


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/consultar")
def consultar_dni():
    payload = request.get_json(silent=True) or {}
    dni = str(payload.get("dni", "")).strip()

    if not _validar_dni(dni):
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "El DNI debe tener exactamente 8 digitos numericos.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    try:
        resultado = client.consultar_dni_playwright(dni)
        return jsonify({"ok": True, **resultado}), HTTPStatus.OK
    except CaptchaRequiredError as exc:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": str(exc),
                    "code": "captcha_required",
                }
            ),
            HTTPStatus.CONFLICT,
        )
    except OnpeLookupError as exc:
        msg = str(exc)
        msg_l = msg.lower()
        if (
            "500" in msg
            or "interno del servidor" in msg_l
            or "pagina no encontrada" in msg_l
            or "página no encontrada" in msg_l
        ):
            msg = (
                "ONPE devolvio un error interno temporal. Intenta nuevamente en unos segundos "
                "o resuelve captcha si aparece en la ventana del navegador."
            )
        return (
            jsonify(
                {
                    "ok": False,
                    "error": msg,
                    "code": "lookup_error",
                }
            ),
            HTTPStatus.BAD_GATEWAY,
        )
    except Exception:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Ocurrio un error inesperado durante la consulta.",
                    "code": "unexpected_error",
                }
            ),
            HTTPStatus.INTERNAL_SERVER_ERROR,
        )


def _extraer_dnis_desde_texto(texto: str) -> list[str]:
    encontrados = re.findall(r"\b\d{8}\b", texto)
    # Mantiene orden de aparicion y elimina duplicados.
    return list(dict.fromkeys(encontrados))


def _procesar_lote_desde_contenido(contenido: str) -> tuple[list[str], list[dict[str, object]]]:
    dnis = _extraer_dnis_desde_texto(contenido)
    resultados: list[dict[str, object]] = []

    if dnis:
        try:
            client.preparar_sesion_lote()
        except CaptchaRequiredError as exc:
            # Si no se resuelve captcha en esta fase, continuamos igual y
            # cada DNI devolvera detalle explicito para el usuario.
            resultados.append(
                {
                    "dni": dnis[0],
                    "ok": False,
                    "error": str(exc),
                    "code": "captcha_required",
                }
            )

    for dni in dnis:
        if any(item.get("dni") == dni for item in resultados):
            continue
        try:
            data = client.consultar_dni_playwright(dni)
            resultados.append({"dni": dni, "ok": True, **data})
        except CaptchaRequiredError as exc:
            resultados.append(
                {
                    "dni": dni,
                    "ok": False,
                    "error": str(exc),
                    "code": "captcha_required",
                }
            )
        except OnpeLookupError as exc:
            resultados.append(
                {
                    "dni": dni,
                    "ok": False,
                    "error": str(exc),
                    "code": "lookup_error",
                }
            )
        except Exception:
            resultados.append(
                {
                    "dni": dni,
                    "ok": False,
                    "error": "Ocurrio un error inesperado durante la consulta.",
                    "code": "unexpected_error",
                }
            )

    return dnis, resultados


def _crear_excel_lote(resultados: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("No se pudo crear la hoja de Excel.")
    sheet.title = "Resultados"

    headers = [
        "DNI",
        "Estado",
        "Nombre",
        "Miembro de mesa",
        "Region / Provincia / Distrito",
        "Ubicacion del local",
        "Local de votacion",
        "Referencia",
        "Rol de mesa",
        "N° de Mesa",
        "N° de Orden",
        "Pabellon",
        "Piso",
        "Aula",
        "Codigo de error",
        "Mensaje",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="0B4F6C")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, item in enumerate(resultados, start=2):
        estado = "OK" if item.get("ok") else "ERROR"
        miembro = item.get("es_miembro_mesa")
        miembro_texto = "SI" if miembro is True else "NO" if miembro is False else "NO DISPONIBLE"

        values = [
            item.get("dni"),
            estado,
            item.get("nombre") or "No disponible",
            miembro_texto,
            item.get("region_provincia_distrito") or "No disponible",
            item.get("ubicacion_local") or "No disponible",
            item.get("local_votacion") or "No disponible",
            item.get("referencia_local") or "No disponible",
            item.get("rol_mesa") or "No disponible",
            item.get("numero_mesa") or "No disponible",
            item.get("numero_orden") or "No disponible",
            item.get("pabellon") or "No disponible",
            item.get("piso") or "No disponible",
            item.get("aula") or "No disponible",
            item.get("code") or "",
            item.get("error") or "",
        ]

        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(resultados) + 1}"

    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        values = [sheet.cell(row=row, column=column_index).value for row in range(1, len(resultados) + 2)]
        length = max(len(str(value)) if value is not None else 0 for value in values)
        sheet.column_dimensions[column_letter].width = min(max(length + 2, len(header) + 2, 12), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.post("/consultar-lote")
def consultar_lote():
    archivo = request.files.get("archivo")
    if archivo is None or not archivo.filename:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Debes adjuntar un archivo .txt o .csv con DNIs.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    try:
        contenido = archivo.read().decode("utf-8", errors="ignore")
    except Exception:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "No se pudo leer el archivo enviado.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    dnis, resultados = _procesar_lote_desde_contenido(contenido)
    if not dnis:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "No se encontraron DNIs validos (8 digitos) en el archivo.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    return (
        jsonify(
            {
                "ok": True,
                "total": len(dnis),
                "procesados": len(resultados),
                "resultados": resultados,
            }
        ),
        HTTPStatus.OK,
    )


@app.post("/consultar-lote-excel")
def consultar_lote_excel():
    archivo = request.files.get("archivo")
    if archivo is None or not archivo.filename:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Debes adjuntar un archivo .txt o .csv con DNIs.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    try:
        contenido = archivo.read().decode("utf-8", errors="ignore")
    except Exception:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "No se pudo leer el archivo enviado.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    dnis, resultados = _procesar_lote_desde_contenido(contenido)
    if not dnis:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "No se encontraron DNIs validos (8 digitos) en el archivo.",
                }
            ),
            HTTPStatus.BAD_REQUEST,
        )

    excel = _crear_excel_lote(resultados)
    return send_file(
        excel,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="consulta_onpe_lote.xlsx",
        max_age=0,
    )


@atexit.register
def _shutdown_browser() -> None:
    client.close()


if __name__ == "__main__":
    # Playwright sync no es thread-safe entre hilos; ejecutamos en modo de un solo hilo.
    app.run(debug=True, use_reloader=False, threaded=False, host="0.0.0.0", port=5001)
